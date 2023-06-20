from whatstk import WhatsAppChat
import re
import os
from openpyxl import Workbook
import shutil
import pandas as pd
filepath = 'input/chat_log.txt'
chat = WhatsAppChat.from_source(filepath=filepath, hformat='[%m/%d/%y, %I:%M:%S %p] %name:')
alllogs=chat.df
row=alllogs.loc[30]
#test    3 4 15 25  35 36   39  43    52  56  58   61    74   75   76   79
j=0
curimg=""
imglist=[]
datarow=[]
prow=[]
qrow=[]
item={
    "date":"",
    "user":"",
    "images":""
}
while j<len(alllogs):
    row=alllogs.loc[j]
    ################################ initialize ######################################
    _workorder=""
    date=""
    dispatcher=""
    _company=""
    nte=""
    amountdue=""
    _techname=""
    technumber=""
    paymentmethod=""
    paymentaddress=""
    joblocation=""
    jobstatus=""
    overall=""
    photosq=""
    workorder=""
    date=""
    dispatcher=""
    company=""
    qotedescription=""
    quotedetails=""
    totalprice=""
    photosp=""
    ################################# end initialize ######################################
    if (row.str.find("<attached:")>0).any() :
        end=row.message.find("jpg")
        image=row.message[11:end+3]
        user=row.username
        date=row.date
        if user==curimg:
            imglist[len(imglist)-1]["images"]=imglist[len(imglist)-1]["images"]+","+image
        else:
            item={
                "date":date,
                "user":user,
                "images":image
            }
            imglist.append(item)            
        curimg=user
        # print(date, user,image)
    else:
        if j!=0 and j%5==0:
            step=1
            while step<5:
                for imageitem in imglist:
                    if datarow[len(datarow)-1-step]["Dispatcher"]==imageitem["user"]:
                        datarow[len(datarow)-1-step]["Photos"]=imageitem["images"]
                        imageitem["user"]=""
                step=step+1

        data=row.message
        str2 = data.splitlines() 
        workorder=-1
        company=-1
        quoteflag=0
        index=1
        jlocationflag=0
        jStatusflag=0
        companyflag=0
        amountflag=0
        pmethodflag=0
        techname=""
        signflag={
            "kind":"Quote",
            "WorkOrder":0,
            "Date":0,
            "Dispatcher":0,
            "Company":0,
            "NTE":0,
            "AmountDue":0,
            "Techname":0,
            "Technumber":0,
            "Paymentmethod":0,
            "Paymentaddress":0,
            "Joblocation":0,
            "Jobstatus":0,
            "Overall":0,
            "qDescription":0,
            "qDetails":0,
            "ToatalPrice":0,
        }
        #################################################     Dispatcher   ##########################################################   
        # print("Dispatcher", row.username)
        dispatcher=row.username
        #################################################     Date   ##########################################################   
        # print("Date:",row.date)        
        date=row.date
        for item in str2:
            if item.strip()!="":   
                item=item.replace("*","") 
                # print("-----",item,"-----")   
                        
            #################################################     Quote   ##########################################################   
                if item.find("quote") >-1 and signflag["WorkOrder"]==0:
                    if  item.find(":"):
                        workorder=item[item.find("quote")+5:].strip().replace(":","").strip()
                        if workorder.find("for")==0:
                            workorder=workorder[3:].strip()
                            if workorder.isdigit():
                                signflag["WorkOrder"]=1
                                # print("Work Order:",workorder)
                                _workorder=workorder.strip()
                                continue                     
                if item.find("QUOTE") >-1 or item.find("update") >-1:
                    quoteflag=1
                    continue
                if quoteflag==1  and signflag["WorkOrder"]==0:
                    quoteflag=0
                    workorder=item.strip()
                    if workorder.find("WO # ")>-1:
                        quoteflag=3
                        signflag["WorkOrder"]=1
                        # print("Work Order:",workorder[3:])
                        _workorder=workorder[5:].strip()
                    if workorder.find("wo#")>-1 or workorder.find("WO#")>-1 or workorder.find("Wo#")>-1:
                        quoteflag=3
                        signflag["WorkOrder"]=1
                        # print("Work Order:",workorder[3:])
                        _workorder=workorder[3:].strip()
                    else:
                        signflag["WorkOrder"]=1
                        # print("Work Order:",workorder)
                        workorder=workorder.replace(" ","")
                        if workorder.find("WO#")>-1:
                            workorder=workorder[3:]
                        _workorder=workorder
                    continue
                if item.find("Quote") >-1:
                    if  item.find("#")>-1:
                        workorder=item[item.find("#")+1:].strip()
                        quoteflag=1
                        signflag["WorkOrder"]=1
                        # print("Work Order:",workorder)
                        _workorder=workorder.strip()
                        continue 
                    else:
                        workorder=item[item.find("Quote")+5:]
                        if workorder.find("for")==0:
                            workorder=workorder[3:].strip().replace(" ","")
                            if workorder.find("WO#")>-1 or workorder.find("Wo#")>-1 or workorder.find("wo#")>-1 :
                                _workorder=workorder[3:].strip()
                            if workorder.isdigit():
                                _workorder=workorder.strip()
                            signflag["WorkOrder"]=1
                            continue
                        # workorder=workorder.strip()
                        # quoteflag=2
                        # signflag["WorkOrder"]=1
                        # # print("Work Order:",workorder)
                        # _workorder=workorder
                        # continue 
                if  quoteflag==1 :
                    quoteflag=0
                    item=item.strip().replace(":","").strip()
                    if item.find("for")==0:
                        item=item[3:].strip().replace(":","").strip()
                        signflag["Company"]=1
                        # print("Company:",item)
                        _company=item.strip()
                        continue
                    else:
                        item=item.strip().replace(":","").strip()
                        signflag["Company"]=1
                        # print("Company:",item)
                        _company=item.strip()
                        continue
                if quoteflag==2:
                    if item.find("for")==0:
                        item=item[3:].strip().replace(":","").strip()
                        if item.find("("):
                            item=item[:item.find("(")]
                            signflag["Company"]=1
                            # print("Company:",item)
                            _company=item.strip()
                            continue
                if quoteflag==3 and item.find("wo#")==-1:
                    item=item.strip().replace(":","").strip()
                    quoteflag=0
                    signflag["Company"]=1
                    # print("Company:",item)
                    _company=item
                    continue
            #################################################     Work Order and Company   ##########################################################     
                if quoteflag==5 and item.find(".com")==-1:
                    item=item.strip().replace(":","").strip()
                    quoteflag=0
                    signflag["Company"]=1
                    # print("Company:",item)
                    _company=item
                    continue
                if workorder==-1:
                    workorder=item.find("WO")
                    if workorder>-1 and item.find("#"):
                        quoteflag=5
                        workorder=item[item.find("#")+1:]
                        signflag["WorkOrder"]=1
                        # print("Work Order:",workorder)
                        _workorder=workorder.strip()
                        continue
                    workorder=item.find("wo")
                    if workorder>-1 and item.find("#"):
                        workorder=item[item.find("#")+1:]
                        signflag["WorkOrder"]=1
                        # print("Work Order:",workorder)
                        companyflag=1  
                        _workorder=workorder.strip()
                        continue
                if type(workorder)!=int:                             
                    if item.find(".com")>-1:  
                        companyflag=1                    
                        continue
                if companyflag==1 and signflag["Company"]==0:
                    companyflag=0
                    item=item.strip().replace(":","").strip()
                    company=item
                    if item.find("description")==-1:
                        # print("Company:",company)
                        _company=company
                        signflag["Company"]=1
                        continue
            #################################################     NTE    ##########################################################
                if item.find("NTE")>-1:
                    item=item[item.find("NTE")+3:]
                    item=item.strip().replace(":","").strip()
                    signflag["NTE"]=1
                    # print("NTE:",item)
                    nte=item
                    continue
            #################################################     Amount Due    ##########################################################
                if item.find("Amount Due")>-1:
                    item=item[item.find("Amount Due")+10:]
                    item=item.strip().replace(":","").strip()
                    signflag["AmountDue"]=1
                    # print("Amount Due:",item)
                    amountdue=item
                    amountflag=1
                    continue
                elif item.find("Ammount Due")>-1:
                    item=item[item.find("Amount Due")+12:]
                    item=item.strip().replace(":","").strip()
                    signflag["AmountDue"]=1
                    # print("Amount Due:",item)
                    amountdue=item
                    amountflag=1
                    continue
            #################################################     Tech name    ##########################################################
                if item.find("Tech full name")>-1:
                    item=item[item.find("Tech full name")+14:]
                    item=item.strip().replace(":","").strip()
                    techname=item
                    signflag["Techname"]=1
                    # print("Tech name:",techname)
                    _techname=techname
                    continue
                elif amountflag==1 and techname=="":
                    amountflag=0
                    item=item.strip().replace(":","").strip()
                    signflag["Techname"]=1
                    # print("Tech name:",item)
                    _techname=item
                    continue
            #################################################     Tech number    ##########################################################
                if item.find("(")>-1 and item.find(")")>-1:
                    sidx=item.find("(")
                    eidx=item.find(")")
                    midx=item.find("-")
                    if eidx>sidx:
                        itemtmp=item[1:eidx-1]+item[eidx+1:midx-1]+item[midx+1:]
                        if re.match("^ *[0-9][0-9 ]*$", itemtmp):
                            Technumber=item
                            signflag["Technumber"]=1
                            # print("Tech number:",Technumber) 
                            technumber=item
                            continue 
                if re.match("^ *[0-9][0-9 ]*$", item):
                    Technumber=item
                    signflag["Technumber"]=1
                    # print("Tech number:",Technumber) 
                    technumber=item
                    continue
                elif item.find("Tech Number")>-1:
                    item=item[item.find("Tech Number")+11:]
                    item=item.strip().replace(":","").strip()
                    signflag["Technumber"]=1
                    technumber=item
                    # print("Tech number:",item)
                    continue
            #################################################     Payment method    ##########################################################
                if item.find("Payment Method")>-1:
                    item=item[item.find("Payment Method")+14:]
                    item=item.strip().replace(":","").strip()
                    signflag["Paymentmethod"]=1
                    paymentmethod=item
                    # print("Payment Method:",item)                    
                    if item.find("Zelle")>-1 or item.find("zelle")>-1 :
                        pmethodflag=2
                        continue
                    elif item.find("Cashapp")>-1:
                        pmethodflag=3
                        continue
                    else:
                        pmethodflag=1
                        continue
                if pmethodflag==2:
                    pmethodflag=0
                    item=item.strip().replace(":","").strip()
                    if item.find("@")>-1:
                        signflag["Paymentmethod"]=1
                        paymentaddress=item
                    else:
                        paymentaddress=technumber
                if pmethodflag==3:
                    pmethodflag=0
                    item=item.strip().replace(":","").strip()
                    if item.find("$")==0:
                        signflag["Paymentmethod"]=1
                        paymentaddress=item[1:]
                if pmethodflag==1:
                    pmethodflag=0
                    item=item.strip().replace(":","").strip()
                    if item.find("@")>-1:
                        signflag["Paymentmethod"]=1
                        paymentaddress=item
                        # print("Payment Address:",item)
            #################################################     Job Location    ##########################################################
                if item.find("Service Location")>-1:
                    item=item[item.find("Service Location")+16:]
                    item=item.strip().replace(":","").strip()
                    signflag["Joblocation"]=1
                    joblocation=item
                    # print("Job Location:",item)
                    continue
                elif item.find("Service location")>-1:
                    item=item[item.find("Service location")+16:]
                    item=item.strip().replace(":","").strip()
                    if item!="":
                        signflag["Joblocation"]=1
                        # print("Job Location:",item)
                        joblocation=item
                        continue
                    else:
                        jlocationflag=1
                        continue
                elif item.find("Job Location")>-1:
                    item=item[item.find("Job Location")+12:]
                    item=item.strip().replace(":","").strip()
                    if item!="":
                        signflag["Joblocation"]=1
                        joblocation=item
                        # print("Job Location:",item)
                        continue
                    else:
                        jlocationflag=1
                        continue
                if jlocationflag==1:
                    item=item.strip().replace(":","").strip()
                    jlocationflag=0
                    signflag["Joblocation"]=1
                    joblocation=item
                    # print("Job Location:",item)
                    continue
            #################################################     Job status    ##########################################################
                if item.find("Job status")>-1 and signflag["Jobstatus"]!=1:
                    item=item[item.find("Job status")+10:]
                    item=item.strip().replace(":","").strip()
                    if item!="":
                        signflag["Jobstatus"]=1
                        # print("Job status:",item)
                        jobstatus=item
                        continue
                    else:
                        jStatusflag=1
                        continue
                elif item.find("Assessment")>-1  and signflag["Jobstatus"]!=1:
                    signflag["Jobstatus"]=1
                    jobstatus=item
                    # print("Job status:",item)
                if jStatusflag==1  and signflag["Jobstatus"]!=1:
                    item=item.strip().replace(":","").strip()
                    jStatusflag=0
                    signflag["Jobstatus"]=1
                    jobstatus=item
                    # print("Job status:",item)
            #################################################     Overall    ##########################################################
                if item.find("Overall")>-1:
                    item=item[item.find("Overall")+7:]
                    item=item.strip().replace(":","").strip()
                    item=item[item.find("$"):]
                    signflag["Overall"]=1
                    # print("Overall:",item)
                    overall=item
                    continue
            #################################################     Total    ##########################################################
                if item.find("Total")>-1:
                    item=item[item.find("Total")+5:]
                    item=item.strip().replace(":","").strip()
                    if len(item.split())==1:
                        signflag["TotalPrice"]=1
                        # print("Total Price:",item)
                        totalprice=item
                        continue
                elif item.find("Grand total")>-1:
                    item=item[item.find("Grand total")+11:]
                    item=item.strip().replace(":","").strip()
                    signflag["TotalPrice"]=1
                    # print("Total Price:",item)
                    totalprice=item
                    continue
                elif item.find("Grand Total")>-1:
                    item=item[item.find("Grand Total")+11:]
                    item=item.strip().replace(":","").strip()
                    signflag["TotalPrice"]=1
                    # print("Total Price:",item)
                    totalprice=item
                    continue
                elif item.find("GRAND TOTAL")>-1:
                    item=item[item.find("GRAND TOTAL")+11:]
                    item=item.strip().replace(":","").strip()
                    signflag["TotalPrice"]=1
                    # print("Total Price:",item)
                    totalprice=item
                    continue
            #################################################     Quote Description    ##########################################################  
                if signflag["qDetails"]==1:
                    quotedetails=item
                    signflag["qDetails"]=0    
                    continue
                if signflag["qDescription"]==1:
                    qotedescription=qotedescription+'\n'+item                    
                    continue
                if item.find("Description :")>-1 or item.find("job description")>-1:
                    signflag["kind"]="Quote"
                    signflag["qDescription"]=1
                    continue
            else:
                if signflag["WorkOrder"]==1 and signflag["qDescription"]==0:
                    signflag["qDescription"]=1
                    continue
                elif signflag["qDescription"]==1:
                    signflag["qDescription"]=0
                    signflag["qDetails"]=1
                    continue
                
                # if quote_descriptionflag and signflag["WorkOrder"]==1:
                #     quote_description=quote_description+item
                # if quote_detailflag and quote_descriptionflag and signflag["WorkOrder"]==1:
                #     quote_detail=item
                #     quote_detailflag=False
                #     quotedetails=quote_detail
                #     quote_detail=""
            # else:
            #     if not quote_descriptionflag and signflag["WorkOrder"]==1:
            #         quote_descriptionflag=True
            #     else:
            #         quote_detailflag=True
            #         quote_descriptionflag=False
            #         qotedescription=quote_description
            #         quote_description=""
            index=index+1        
        data_row={
            "WorkOrder":_workorder,
            "Date":date,
            "Dispatcher":dispatcher,
            "Company":_company,
            "NTE":nte,
            "AmountDue":amountdue,
            "Techname":_techname,
            "Technumber":technumber,
            "Paymentmethod":paymentmethod,
            "PaymentAddress":paymentaddress,
            "JobLocation":joblocation,
            "JobStatus":jobstatus,
            "Overall":overall,
            "Photos":"",
            "QuoteDescription":qotedescription,
            "QuoteDetails":quotedetails,
            "TotalPrice":totalprice,
            "kind":""
        }
        if signflag["Paymentmethod"]==1 or  signflag["Paymentaddress"]==1  or  signflag["Techname"]==1 or signflag["Technumber"]==1      :
            signflag["kind"]="Payment"
            data_row["kind"]="payment"
           
        else:
            signflag["kind"]="Quote"
            data_row["kind"]="quote"
        datarow.append(data_row)    
    j=j+1 
step=1
while step<5:
    for imageitem in imglist:
        if datarow[len(datarow)-1-step]["Dispatcher"]==imageitem["user"]:
            datarow[len(datarow)-1-step]["Photos"]=imageitem["images"]
            imageitem["user"]=""
    step=step+1
os.mkdir("output")
df = pd.DataFrame(datarow)  
out_quote=[]
out_payment=[]
for item in datarow:
    if item["WorkOrder"]!="":
        if item["kind"]=="quote":
            out_quote.append(item)
            if item["Photos"]!="":
                srcimgs=item["Photos"].split(",")
                dist="output/"+item["WorkOrder"].strip()+"_"+item["Date"].strftime("%m_%d_%Y_%H_%M_%S")+"-quote"
                os.mkdir(dist)
                for srcs in srcimgs:
                    src='input/'+srcs.strip()
                    shutil.copy(src,dist+"/")
        if item["kind"]=="payment":
            out_payment.append(item)
            if item["Photos"]!="":
                srcimgs=item["Photos"].split(",")
                dist="output/"+item["WorkOrder"].strip()+"_"+item["Date"].strftime("%m_%d_%Y_%H_%M_%S")+"-payment"
                os.mkdir(dist)
                for srcs in srcimgs:
                    src='input/'+srcs.strip()
                    shutil.copy(src,dist+"/")
df_quote=pd.DataFrame(out_quote)  
df_payment=pd.DataFrame(out_payment) 
df1=df_quote[["WorkOrder","Date","Dispatcher","Company","NTE","AmountDue","Techname","Technumber", "Paymentmethod","PaymentAddress","JobLocation","JobStatus","Overall","Photos"]]
df2=df_payment[["WorkOrder","Date","Dispatcher","Company","QuoteDescription","QuoteDetails","TotalPrice","Photos"]]


# df2.to_excel('output_quote.xlsx', index=False)
# df1.to_excel('output_payment.xlsx', index=False)


wb = Workbook()
sheet = wb.active
k=0
sheet.cell(row=1, column=1).value = "WorkOrder"
sheet.cell(row=1, column=2).value = "Date"
sheet.cell(row=1, column=3).value = "Dispatcher"
sheet.cell(row=1, column=4).value = "Company"
sheet.cell(row=1, column=5).value = "QuoteDescription"
sheet.cell(row=1, column=6).value = "QuoteDetails"
sheet.cell(row=1, column=7).value = "TotalPrice"
while k<len(out_quote):
    if out_quote[k]["Photos"]!="":
        link="output/"+out_quote[k]["WorkOrder"].strip()+"_"+out_quote[k]["Date"].strftime("%m_%d_%Y_%H_%M_%S")+"-quote/"
        sheet.cell(row=k+2, column=1).hyperlink = link
        sheet.cell(row=k+2, column=1).style = "Hyperlink"
    sheet.cell(row=k+2, column=1).value = out_quote[k]["WorkOrder"]
    sheet.cell(row=k+2, column=2).value = out_quote[k]["Date"]
    sheet.cell(row=k+2, column=3).value = out_quote[k]["Dispatcher"]
    sheet.cell(row=k+2, column=4).value = out_quote[k]["Company"]
    sheet.cell(row=k+2, column=5).value = out_quote[k]["QuoteDescription"]
    sheet.cell(row=k+2, column=6).value = out_quote[k]["QuoteDetails"]
    sheet.cell(row=k+2, column=7).value = out_quote[k]["TotalPrice"]
    k=k+1
wb.save("output_quote.xlsx")

wb1 = Workbook()
sheet1 = wb1.active
k=0
sheet1.cell(row=1, column=1).value = "WorkOrder"
sheet1.cell(row=1, column=2).value = "Date"
sheet1.cell(row=1, column=3).value = "Dispatcher"
sheet1.cell(row=1, column=4).value = "Company"
sheet1.cell(row=1, column=5).value = "NTE"
sheet1.cell(row=1, column=6).value = "AmountDue"
sheet1.cell(row=1, column=7).value = "Techname"
sheet1.cell(row=1, column=8).value = "Technumber"
sheet1.cell(row=1, column=9).value = "Paymentmethod"
sheet1.cell(row=1, column=10).value = "PaymentAddress"
sheet1.cell(row=1, column=11).value = "JobLocation"
sheet1.cell(row=1, column=12).value = "JobStatus"
sheet1.cell(row=1, column=13).value = "Overall"
while k<len(out_payment):
    if out_payment[k]["Photos"]!="":
        link="output/"+out_payment[k]["WorkOrder"].strip()+"_"+out_payment[k]["Date"].strftime("%m_%d_%Y_%H_%M_%S")+"-payment"
        sheet1.cell(row=k+2, column=1).hyperlink = link
        sheet1.cell(row=k+2, column=1).style = "Hyperlink"
    sheet1.cell(row=k+2, column=1).value = out_payment[k]["WorkOrder"]
    sheet1.cell(row=k+2, column=2).value = out_payment[k]["Date"]
    sheet1.cell(row=k+2, column=3).value = out_payment[k]["Dispatcher"]
    sheet1.cell(row=k+2, column=4).value = out_payment[k]["Company"]
    sheet1.cell(row=k+2, column=5).value = out_payment[k]["NTE"]
    sheet1.cell(row=k+2, column=6).value = out_payment[k]["AmountDue"]
    sheet1.cell(row=k+2, column=7).value = out_payment[k]["Techname"]
    sheet1.cell(row=k+2, column=8).value = out_payment[k]["Technumber"]
    sheet1.cell(row=k+2, column=9).value = out_payment[k]["Paymentmethod"]
    sheet1.cell(row=k+2, column=10).value = out_payment[k]["PaymentAddress"]
    sheet1.cell(row=k+2, column=11).value = out_payment[k]["JobLocation"]
    sheet1.cell(row=k+2, column=12).value = out_payment[k]["JobStatus"]
    sheet1.cell(row=k+2, column=13).value = out_payment[k]["Overall"]
    k=k+1
wb1.save("output_payment.xlsx")
